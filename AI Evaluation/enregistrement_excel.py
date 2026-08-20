import openpyxl as oxl
import numpy as np
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter


from Evaluation_IA import Liste_IA, Nb_pb, Nb_crit, Criteres, categories_names, langues, alpha, beta, delta, eta


def write_line(wb,wordsheet,l_values,start_row=1,start_col=1):
    """
    Description : 
    Ecrit dans le workbook wb, sur la page wordsheet, les valeurs de l_values en commençant par la case (start_row, start_col) et en continuant horizontalement
    
    Entrées : 
    - wb : openpyxl workbook
    - wordsheet : str
    - l_values : list of str
    - start_row : int
    - start_ : int
    
    Sorties : Aucune
    """
    ws = wb[wordsheet]
    # ws.append(l_values)
    for i in range(len(l_values)):
        ws.cell(row=start_row, column=start_col + i, value=l_values[i])

        

def add_border_side(cell, left=None, right=None, top=None, bottom=None):
    """
    Ajoute des côtés de bordure à une cellule sans écraser les autres.
    """

    old = cell.border

    cell.border = Border(
        left=left or old.left,
        right=right or old.right,
        top=top or old.top,
        bottom=bottom or old.bottom)
    
    
    
def dessine_cadre_bordures(ws, x1, x2, y1, y2):
    """
    
    """
    thin = Side(style="thin")
    
    for i in range(x1,x2+1):
        add_border_side(ws.cell(row=y1, column=i), top = thin)
        add_border_side(ws.cell(row=y2, column=i), bottom = thin)
    for i in range(y1,y2+1):
        add_border_side(ws.cell(row=i, column=x1), left = thin)
        add_border_side(ws.cell(row=i, column=x2), right = thin)
  
    

def initialisation_etude():
    """
    Description : 
    Prépare le tableur dans lequel seront enregistrées toutes les mesures. 
    Il possède une page pour chaque modèle d'IA évalué, et une page de comparaison
    
    Entrées : Aucune
    
    Sorties : 
    - Renvoie le workbook associé au tableur
    """
    global Liste_IA, Nb_pb, Criteres, langues, Nb_crit
    global alpha, beta, delta, eta

    nb_ia = len(Liste_IA)
    
    # création du workbook
    wb = oxl.Workbook()
    ws_main = wb.active
    ws_main.title = "Comparatif"
    
    
    Nb_crit_with_tot = Nb_crit+1
    taille_col = 22
    taille_col_ia = 23
    
    
              
    # Remplissage des entêtes des sous-tableaux de la Comparatif
    for l in range(len(langues)):
        
        # Case langue
        ws_main.merge_cells(f'A{1+l*(nb_ia + 5)}:C{1+l*(nb_ia + 5)}')
        ws_main[f'A{1+l*(nb_ia + 5)}'] = langues[l]
        ws_main[f'A{1+l*(nb_ia + 5)}'].alignment = oxl.styles.Alignment(horizontal = 'center')
        ws_main[f'A{1+l*(nb_ia + 5)}'].font = oxl.styles.Font(size = 14, bold = True)


        # Pour chaque sous-tableau
        for cat in range(len(categories_names)):
            # Case titre du tableau
            ws_main.merge_cells(f'{get_column_letter(2+(Nb_crit_with_tot+3)*cat)}{2+l*(nb_ia+5)}:{get_column_letter(1+(Nb_crit_with_tot+3)*cat+Nb_crit_with_tot)}{2+l*(nb_ia+5)}')
            ws_main[f'{get_column_letter(2+(Nb_crit_with_tot+3)*cat)}{2+l*(nb_ia+5)}'] = categories_names[cat]
            ws_main[f'{get_column_letter(2+(Nb_crit_with_tot+3)*cat)}{2+l*(nb_ia+5)}'].alignment = oxl.styles.Alignment(horizontal = 'center')
            ws_main[f'{get_column_letter(2+(Nb_crit_with_tot+3)*cat)}{2+l*(nb_ia+5)}'].font = oxl.styles.Font(size = 12, bold = True)

            write_line(wb,"Comparatif",
                        ["Nom IA","Longueur de réponse","Nombre de mots","Densité d'information utile","Note de réussite","Score"],
                        start_row=3+l*(nb_ia+5),
                        start_col=1+cat*(3+Nb_crit_with_tot))
    
            
            # Ajout des bordures du tableau
            dessine_cadre_bordures(ws_main,
                                   x1 = 2+(Nb_crit_with_tot+3)*cat,
                                   x2 = 1+(Nb_crit_with_tot+3)*cat + Nb_crit_with_tot,
                                   y1 = 2+l*(nb_ia+5),
                                   y2 = 3+l*(nb_ia+5) + nb_ia)
            dessine_cadre_bordures(ws_main,
                                   x1 = 1+(Nb_crit_with_tot+3)*cat,
                                   x2 = 1+(Nb_crit_with_tot+3)*cat + Nb_crit_with_tot,
                                   y1 = 3+l*(nb_ia+5),
                                   y2 = 3+l*(nb_ia+5) + nb_ia)
            dessine_cadre_bordures(ws_main,
                                   x1 = 1+(Nb_crit_with_tot+3)*cat,
                                   x2 = 1+(Nb_crit_with_tot+3)*cat + Nb_crit_with_tot,
                                   y1 = 4+l*(nb_ia+5),
                                   y2 = 3+l*(nb_ia+5) + nb_ia)
            dessine_cadre_bordures(ws_main,
                                   x1 = 1+(Nb_crit_with_tot+3)*cat + Nb_crit_with_tot,
                                   x2 = 1+(Nb_crit_with_tot+3)*cat + Nb_crit_with_tot,
                                   y1 = 3+l*(nb_ia+5),
                                   y2 = 3+l*(nb_ia+5) + nb_ia)
    
    
    
    
    
    
    
    
    # ---------
    # Initialisation des pages dédiées aux IA
    # ---------
    
    for i in range(nb_ia):
        # Création de la page de l'IA
        ws = wb.create_sheet(Liste_IA[i])
        
        # Remplissage des entêtes de chaque sous-tableau (langue, mode d'utilisation)
        for l in range(len(langues)):
            ws.merge_cells(f'A{1+l*(Nb_pb + 5)}:C{1+l*(Nb_pb + 5)}')
            ws[f'A{1+l*(Nb_pb + 5)}'] = langues[l]
            ws[f'A{1+l*(Nb_pb + 5)}'].alignment = oxl.styles.Alignment(horizontal = 'center')
            ws[f'A{1+l*(Nb_pb + 5)}'].font = oxl.styles.Font(size = 14, bold = True)
            for cat in range(len(categories_names)):    
                ws.merge_cells(f'{get_column_letter(2+(Nb_crit+3)*cat)}{2+l*(Nb_pb + 5)}:{get_column_letter(1+(Nb_crit+3)*cat+Nb_crit)}{2+l*(Nb_pb + 5)}')
                ws[f'{get_column_letter(2+(Nb_crit+3)*cat)}{2+l*(Nb_pb + 5)}'] = categories_names[cat]
                ws[f'{get_column_letter(2+(Nb_crit+3)*cat)}{2+l*(Nb_pb + 5)}'].alignment = oxl.styles.Alignment(horizontal = 'center')
                ws[f'{get_column_letter(2+(Nb_crit+3)*cat)}{2+l*(Nb_pb + 5)}'].font = oxl.styles.Font(size = 12, bold = True)
                write_line(wb,Liste_IA[i],
                        ["N° Problème","Longueur de réponse","Nombre de mots","Densité d'information utile","Note de réussite"],
                        start_row=3+l*(Nb_pb+5),
                        start_col=1+cat*(3+Nb_crit))
            
            
            
                # Tracé des bordures de chaque sous-tableau (langue, mode d'utilisation)
                dessine_cadre_bordures(ws,
                        x1 = 1+(Nb_crit+3)*cat,
                        x2 = 1+(Nb_crit+3)*cat + Nb_crit,
                        y1 = 3+l*(Nb_pb+5),
                        y2 = 3+l*(Nb_pb+5)+Nb_pb)
                dessine_cadre_bordures(ws,
                        x1 = 2+(Nb_crit+3)*cat,
                        x2 = 1+(Nb_crit+3)*cat + Nb_crit,
                        y1 = 2+l*(Nb_pb+5),
                        y2 = 3+l*(Nb_pb+5)+Nb_pb)
                dessine_cadre_bordures(ws,
                        x1 = 1+(Nb_crit+3)*cat,
                        x2 = 1+(Nb_crit+3)*cat + Nb_crit,
                        y1 = 3+l*(Nb_pb+5),
                        y2 = 3+l*(Nb_pb+5))
        
        
            # Ajout de la ligne bilan de l'IA en cours dans la page principale
            for cat in range(len(categories_names)):
                write_line(wb, "Comparatif",
    [
        Liste_IA[i],
        f"=AVERAGE('{Liste_IA[i]}'!{get_column_letter(2+(Nb_crit+3)*cat)}{4+l*(Nb_pb + 5)}:{get_column_letter(2+(Nb_crit+3)*cat)}{Nb_pb+3+l*(Nb_pb + 5)})",
        f"=AVERAGE('{Liste_IA[i]}'!{get_column_letter(2+(Nb_crit+3)*cat+1)}{4+l*(Nb_pb + 5)}:{get_column_letter(2+(Nb_crit+3)*cat+1)}{Nb_pb+3+l*(Nb_pb + 5)})",
        f"=AVERAGE('{Liste_IA[i]}'!{get_column_letter(2+(Nb_crit+3)*cat+2)}{4+l*(Nb_pb + 5)}:{get_column_letter(2+(Nb_crit+3)*cat+2)}{Nb_pb+3+l*(Nb_pb + 5)})",
        f"=AVERAGE('{Liste_IA[i]}'!{get_column_letter(2+(Nb_crit+3)*cat+3)}{4+l*(Nb_pb + 5)}:{get_column_letter(2+(Nb_crit+3)*cat+3)}{Nb_pb+3+l*(Nb_pb + 5)})"
    ],
    start_row=4+i+l*(nb_ia+5),
    start_col=1+(Nb_crit_with_tot+3)*cat
)
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = taille_col_ia    


    # ---------
    # Ajout de la formule de calcul du score
    # ---------
    
    # Liste des cases dans l'angle supérieur droit des tableaux de la page Comparatif
    cases_long_carac_deb = [(2+cat*(3+Nb_crit_with_tot),4+l*(5+nb_ia))
            for l in range(len(langues))
            for cat in range(len(categories_names))
            ]

    # Liste des cases score de la page Comparatif
    cases_score_deb = translation(cases_long_carac_deb,4,0) 
    cases_score = [
    (c[0], c[1] + k)
    for c in cases_score_deb
    for k in range(nb_ia)
]

    # Ajout de la formule de calcul du score
    for c in cases_score:
        ws_main[f'{get_column_letter(c[0])}{c[1]}'] = (
          f'= ({get_column_letter(c[0]-2)}{c[1]}^{alpha} *'
          f' ({get_column_letter(c[0]-1)}{c[1]}/10)^{beta} ) /'
          f' (EXP({delta} *'
          f' MAX({get_column_letter(c[0]-3)}{c[1]}/{eta}-1,0)))'
      )
    
    # ---------
    # Coloration des cases de la page principale 
    # pour faire ressortir les critères les plus importants
    # ---------
    red_color = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_color = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    
    for l in range(len(langues)):
        for cat in range(len(categories_names)):
            for j in range(1,Nb_crit_with_tot+1):
                col_letter = get_column_letter(j+cat*(Nb_crit_with_tot+3)+1)
            
                range_str = f"{col_letter}{4+l*(nb_ia+5)}:{col_letter}{nb_ia+3+l*(nb_ia+5)}"
                
                formula_max = f"{col_letter}{4+l*(nb_ia+5)}=MAX(${col_letter}${4+l*(nb_ia+5)}:${col_letter}{nb_ia+3+l*(nb_ia+5)})"
                formula_min = f"{col_letter}{4+l*(nb_ia+5)}=MIN(${col_letter}${4+l*(nb_ia+5)}:${col_letter}{nb_ia+3+l*(nb_ia+5)})"
                
                if Criteres[j-1]["Objectif"] == "Minimiser":
                    ws_main.conditional_formatting.add(
                        range_str,
                        FormulaRule(formula=[formula_max], fill=red_color)
                    )
                    ws_main.conditional_formatting.add(
                        range_str,
                        FormulaRule(formula=[formula_min], fill=green_color)
                    )
                
                elif Criteres[j-1]["Objectif"] == "Maximiser":
                    ws_main.conditional_formatting.add(
                        range_str,
                        FormulaRule(formula=[formula_max], fill=green_color)
                    )
                    ws_main.conditional_formatting.add(
                        range_str,
                        FormulaRule(formula=[formula_min], fill=red_color)
                    )
                
    # Elargissement des colonnes de la page principale     
    for col in ws_main.columns:
        col_letter = col[3].column_letter
        ws_main.column_dimensions[col_letter].width = taille_col      
    
    return wb





def save(wb,filename):
    """
    Description : 
    Enregistre le workbook wb sous le nom filename
    
    Entrées : 
    - wb : openpyxl workbook
    - filename : str
    
    Sorties : Aucune
    """
    wb.save(filename)
    



def translation(l,x,y):
    """
    Description : 
    Renvoie la liste l translatée de (x,y)
    
    Entrées : 
    - l : list of tuple of int
    - x : int
    - y : int
    
    Sorties :
    - l_translated : list of tuple of int
    """
    l_translated = []
    for t in l:
        l_translated.append((t[0]+x, t[1]+y))
    return l_translated


def range_2_à_2(l1,l2):
    """
    Description : 
    Renvoie une chaîne de caractères représentant (en langage Excel) 
    l'union des ensembles de cases situées entre (x1,y1) et (x2,y2), 
    les éléments respectifs de l1 et l2, de même indice
    
    Condition : l1 et l2 doivent être de même longueur
    
    Entrées : 
    - l1 : list of tuple of int
    - l2 : list of tuple of int

    
    Sorties :
    - s : str
    """
    assert(len(l1) == len(l2))
    
    s = ""
    for i in range(len(l1)):
        s+=f"{get_column_letter(l1[i][0])}{l1[i][1]}:{get_column_letter(l2[i][0])}{l2[i][1]};"
        
    return s[0:-1]





def copie_notes(wb):
    """
    Description : 
    Copie les notes de chaque problème de l'étude, stockées dans le sheet
    "Notes.xlsx", dans le tableau récapitulatif de l'étude
        
    Entrées : 
    - wb : workbook d'étude

    
    Sorties : Aucune
    """
    # Ouverture du tableau des notes, en n'utilisant que les valeurs (sinon on copie les formules SUM(...))
    wb_notes = oxl.load_workbook("Notes.xlsx", data_only=True)

    
    # Copie des notes pour chaque problème dans chaque configuration
    for ia in Liste_IA:
        for l in range(len(langues)):
            for cat in range(len(categories_names)):
                
                for i in range(Nb_pb):
                    wb[ia].cell(4+i+l*(Nb_pb+5),5+cat*(Nb_crit+3)).value = wb_notes[ia][4+i+l*(Nb_pb+9)][7+cat*11].value
                    


